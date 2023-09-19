import openpyxl


class HomePageData():

    test_HomePage_data = [{"firstname":"Mujeeb","email":"mujeeb.rahaman@cognine.com","password":"Mujeeb@123","gender":"Male","myMessage":"HelloWorld!"},{"firstname":"Alia","email":"alia.bhatt@gmail.com","password":"Alia@123","gender":"Female","myMessage":"MyWorld!!"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Mujeeb Rahaman\\Downloads\\PythonDocs\\Pythondemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):  # To get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # To get columns
                    # Dict["email"]="alia.bhatt@gmail.com"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]
